import  openpyxl
wb= openpyxl.load_workbook('example.xlsx')
# print(wb.sheetnames)
# for sheet in wb:
#     print(sheet.title)
mysheet = wb.create_sheet('my sheet')
print(wb.sheetnames)
